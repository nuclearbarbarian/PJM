"""
Ingest PJM Interconnection Queue Data (Serial + Cycle)
Source: PJM public queue exports (not Data Miner 2)
Outputs: data/processed/queue.json (Ultan-compliant)

Geocodes projects to county centroids using US Census FIPS.
"""

import json
import openpyxl
from datetime import date
from pathlib import Path

RAW_DIR = Path(__file__).parent.parent / "raw"
OUT_DIR = Path(__file__).parent.parent / "processed"
OUT_DIR.mkdir(exist_ok=True)

# ── County centroids (computed from US Census TopoJSON) ──────────
# Loaded from pre-built lookup; falls back to state centroids if county not found
SCRIPTS_DIR = Path(__file__).parent

def load_county_centroids():
    """Load county centroids from pre-built JSON lookup."""
    path = SCRIPTS_DIR / "county_centroids.json"
    if path.exists():
        with open(path) as f:
            return json.load(f)
    return {}

COUNTY_CENTROIDS = load_county_centroids()

# Fallback: state-level centroids for when county lookup fails
STATE_CENTROIDS = {
    "PA": (40.9, -77.8), "NJ": (40.1, -74.7), "DE": (39.0, -75.5),
    "MD": (39.0, -76.7), "VA": (37.5, -79.0), "WV": (38.6, -80.6),
    "OH": (40.4, -82.8), "IN": (39.8, -86.3), "IL": (40.0, -89.2),
    "MI": (44.3, -84.5), "KY": (37.8, -84.3), "NC": (35.6, -79.8),
    "TN": (35.9, -86.4), "DC": (38.9, -77.0),
}


def geocode(state, county):
    """Geocode to county centroid, falling back to state centroid."""
    if county and county not in ("Unknown", "N/A", ""):
        key = f"{state}:{county}"
        if key in COUNTY_CENTROIDS:
            return COUNTY_CENTROIDS[key]
    return STATE_CENTROIDS.get(state, (None, None))


def parse_serial_queue():
    """Parse the serial queue Excel into a list of dicts."""
    wb = openpyxl.load_workbook(RAW_DIR / "pjm_queue_serial.xlsx", read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]

    projects = []
    for row in rows[1:]:
        d = dict(zip(headers, row))

        # Extract submitted year
        submitted = d.get("Submitted Date", "")
        submitted_year = None
        if submitted:
            try:
                parts = str(submitted).split("/")
                if len(parts) == 3:
                    submitted_year = int(parts[2])
            except (ValueError, IndexError):
                pass

        # Parse MW values safely
        def safe_float(val):
            try:
                return float(val) if val is not None and val != "" else None
            except (ValueError, TypeError):
                return None

        state = str(d.get("State", "")).strip()
        county = str(d.get("County", "")).strip()

        # Geocode to county centroid, with state centroid fallback
        lat, lon = geocode(state, county)

        project = {
            "id": d.get("Project ID"),
            "name": d.get("Name"),
            "commercial_name": d.get("Commercial Name"),
            "state": state,
            "county": county,
            "status": d.get("Status"),
            "transmission_owner": d.get("Transmission Owner"),
            "mw_energy": safe_float(d.get("MW Energy")),
            "mw_capacity": safe_float(d.get("MW Capacity")),
            "mw_in_service": safe_float(d.get("MW In Service")),
            "capacity_or_energy": d.get("Capacity or Energy"),
            "project_type": d.get("Project Type"),
            "fuel": d.get("Fuel"),
            "submitted_date": str(submitted) if submitted else None,
            "submitted_year": submitted_year,
            "projected_in_service": str(d.get("Projected In Service Date", "")) or None,
            "actual_in_service": str(d.get("Actual In Service Date", "")) or None,
            "withdrawal_date": str(d.get("Withdrawal Date", "")) or None,
            "lat": lat,
            "lon": lon,
            "queue_type": "serial",
            "geocode_method": "state_centroid" if lat else None,
        }
        projects.append(project)

    wb.close()
    return projects


def parse_cycle_queue():
    """Parse the cycle queue Excel into a list of dicts."""
    wb = openpyxl.load_workbook(RAW_DIR / "pjm_queue_cycle.xlsx", read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]

    projects = []
    for row in rows[1:]:
        d = dict(zip(headers, row))

        submitted = d.get("Submitted Date", "")
        submitted_year = None
        if submitted:
            try:
                parts = str(submitted).split("/")
                if len(parts) == 3:
                    submitted_year = int(parts[2])
            except (ValueError, IndexError):
                pass

        def safe_float(val):
            try:
                return float(val) if val is not None and val != "" else None
            except (ValueError, TypeError):
                return None

        state = str(d.get("State", "")).strip()
        county = str(d.get("County", "")).strip()
        lat, lon = geocode(state, county)

        project = {
            "id": d.get("Project ID"),
            "cycle": d.get("Cycle"),
            "stage": d.get("Stage"),
            "name": d.get("Name"),
            "commercial_name": d.get("Commercial Name"),
            "developer": d.get("Developer"),
            "state": state,
            "county": str(d.get("County", "")).strip(),
            "status": d.get("Status"),
            "transmission_owner": d.get("Transmission Owner"),
            "mw_energy": safe_float(d.get("MW Energy")),
            "mw_capacity": safe_float(d.get("MW Capacity")),
            "mw_in_service": safe_float(d.get("MW In Service")),
            "capacity_or_energy": d.get("Capacity or Energy"),
            "project_type": d.get("Project Type"),
            "fuel": d.get("Fuel"),
            "submitted_date": str(submitted) if submitted else None,
            "submitted_year": submitted_year,
            "lat": lat,
            "lon": lon,
            "queue_type": "cycle",
            "geocode_method": "state_centroid" if lat else None,
        }
        projects.append(project)

    wb.close()
    return projects


def main():
    serial = parse_serial_queue()
    cycle = parse_cycle_queue()
    all_projects = serial + cycle

    # Summary stats
    fuels = {}
    statuses = {}
    years = {}
    for p in all_projects:
        f = p.get("fuel") or "Unknown"
        fuels[f] = fuels.get(f, 0) + 1
        s = p.get("status") or "Unknown"
        statuses[s] = statuses.get(s, 0) + 1
        y = p.get("submitted_year")
        if y and y >= 2020:
            years[y] = years.get(y, 0) + 1

    output = {
        "source_url": "https://www.pjm.com/planning/services-requests/interconnection-queues",
        "source_description": "PJM Interconnection Queue — Serial and Cycle Service Request Status (public Excel exports)",
        "access_date": date.today().isoformat(),
        "jurisdiction": "PJM Interconnection, LLC",
        "citation": "PJM Interconnection, LLC. 'Interconnection Queues.' Accessed {}.".format(
            date.today().strftime("%B %d, %Y")
        ),
        "known_gaps": [
            "No latitude/longitude coordinates — CEII-restricted. Geocoded to state centroids.",
            "County field contains 'Unknown' or 'Not Specified' for some projects.",
            "Serial queue covers all historical projects back to 1997; filtered to 2020+ for dashboard.",
            "Cycle queue covers TC1, TC2, and Cycle 01 only.",
        ],
        "notes": "Data sourced from PJM public Excel exports, not Data Miner 2 API. "
                 "Non-member redistribution restrictions on Data Miner 2 do not apply to these public documents.",
        "summary": {
            "total_projects": len(all_projects),
            "serial_projects": len(serial),
            "cycle_projects": len(cycle),
            "fuel_breakdown": dict(sorted(fuels.items(), key=lambda x: -x[1])),
            "status_breakdown": dict(sorted(statuses.items(), key=lambda x: -x[1])),
            "projects_by_year_2020_plus": dict(sorted(years.items())),
        },
        "data": all_projects,
    }

    out_path = OUT_DIR / "queue.json"
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2, default=str)

    print(f"Wrote {len(all_projects)} projects to {out_path}")
    print(f"  Serial: {len(serial)}, Cycle: {len(cycle)}")
    print(f"  Fuels: {fuels}")
    print(f"  Statuses: {statuses}")
    print(f"  2020+ by year: {years}")


if __name__ == "__main__":
    main()
